import openpyxl
import requests
import pandas as pd
import json
import ast
from collections import MutableMapping


def convert_flatten(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k

        if isinstance(v, MutableMapping):
            items.extend(convert_flatten(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)

a = 4/2
#print(a)
b = 4//2
#print(b)

payload = {'api_key':'API_KEY_REMOVED'}

writer = pd.ExcelWriter('postmaninfo.xlsx')
# Requete GET
resp = requests.get('https://jsonplaceholder.typicode.com/users?$top=1'
                    )
data = resp.text
data = data.replace('[','')
data = data.replace(']','')
data = str(data)
data = json.loads(json.dumps(resp.json()))

df = pd.DataFrame(data)
TotalColumns = list(df.shape)

d = {}
listofcolumns = df.columns.values.tolist()

#print(listofcolumns)
for i in range(0,TotalColumns[1]):
    d[listofcolumns[i]] = df[listofcolumns[i]].values.tolist()

dfbasic = pd.DataFrame(d)
dfbasic = dfbasic.fillna(0)
print(dfbasic.columns)
totalnumberofrows = list(dfbasic.shape)
print('totla number of rows::',str(totalnumberofrows))

df0 = pd.DataFrame()
d3 = {}
l1= []
workbook = openpyxl.load_workbook('json.xlsx')
worksheet = workbook.active
k = 1
j=1
for i in range(0, 10):
    d3['id'] = d['id'][i]
    d3['company'] = d['company'][i]
    d3 = convert_flatten(d3)
    df0 = pd.DataFrame(d3,index=[0])
    print(df0.columns)
    worksheet.cell(k,j).value = df0['id'][0]
    worksheet.cell(k, j+1).value = df0['company_name'][0]
    worksheet.cell(k, j+2).value = df0['company_catchPhrase'][0]
    worksheet.cell(k, j+3).value = df0['company_bs'][0]
    k = k+1
workbook.save('json.xlsx')
dfcmpany = pd.read_excel('json.xlsx')
dfcmpany = pd.DataFrame(dfcmpany)
print(dfcmpany.columns)
dffinal = pd.merge(left=dfbasic,right=dfcmpany,on='id', how='inner')
print(dffinal.shape)
dffinal.to_excel(writer,'company')


#print(pd.DataFrame(l1,index=[0]))






#d3 = {}
#d3['id'] = d['id'][0]
#d3['company'] = d['company'][0]
#print('printing dictories')
#print(d3)
#print(convert_flatten(d3))
#d3 = convert_flatten(d3)
#dfcmpany = pd.DataFrame(d3,index=[0])
#print(dfcmpany)





#dfbasic.to_excel(writer,'basicinfo')

#dfaddress = pd.DataFrame(d['address'])
#print(dfaddress.columns)
#dfaddress = dfaddress.fillna(0)
#dfaddress.to_excel(writer,'address')

#dfaddress = pd.DataFrame(d['address'][1]['geo'])
#dfaddress.to_excel(writer,'geo.xlsx')

#dfcmpany = pd.DataFrame(d['company'])
#print(dfcmpany)
#dfcmpany["id"] = d['id']
#print(dfcmpany.columns)
#dfcmpany = dfcmpany.fillna(0)
#dfcmpany.to_excel(writer,'company')


#dffinal = pd.join(left='dfbasic',right='dfcmpany',on='id',how='inner')
#dffinal.to_excel(writer,'final')
writer.save()




















